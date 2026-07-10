from collections import defaultdict
from datetime import datetime
import calendar
from pathlib import Path
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from services.config_store import load_sop_codes, load_workers
from services.session_store import get_session_exports_directory, load_session


NOTE_HIGHLIGHT_CODES = {"AL", "MC", "BL", "OFF"}
HOURS_PATTERN = re.compile(r"^\s*(?:(\d+)\s*h)?\s*(?:(\d+)\s*m)?\s*$", re.IGNORECASE)
STANDARD_SOP_CODES = ["AL", "MC", "BL", "OFF"]
NOTE_FILL = PatternFill(fill_type="solid", fgColor="FFFF00")
ATTENTION_FILL = PatternFill(fill_type="solid", fgColor="F4B183")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="EAF2F8")
SUMMARY_FILL = PatternFill(fill_type="solid", fgColor="F4F6F7")
LIGHT_ORANGE_FILL = PatternFill(fill_type="solid", fgColor="FCE4D6")
WHITE_FILL = PatternFill(fill_type="solid", fgColor="FFFFFF")
THIN_SIDE = Side(style="thin", color="D9E2EC")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)


def _normalize_note_code(note: object) -> str:
    return str(note or "").strip().upper()


def _parse_hours_to_minutes(hours_worked: object) -> int:
    value = str(hours_worked or "").strip()
    if not value:
        return 0

    match = HOURS_PATTERN.match(value)
    if not match:
        return 0

    hours = int(match.group(1) or 0)
    minutes = int(match.group(2) or 0)
    return (hours * 60) + minutes


def _format_minutes(total_minutes: int) -> str:
    hours, minutes = divmod(total_minutes, 60)
    return f"{hours}h {minutes:02d}m"


def _parse_work_date(value: object) -> datetime | None:
    date_text = str(value or "").strip()
    if not date_text:
        return None

    for date_format in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(date_text, date_format)
        except ValueError:
            continue
    return None


def _safe_sheet_title(worker_name: str, used_titles: set[str]) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]", "", worker_name).strip() or "Worker"
    title = cleaned[:31]
    if title not in used_titles:
        used_titles.add(title)
        return title

    suffix = 2
    while True:
        suffix_text = f" {suffix}"
        candidate = f"{cleaned[:31 - len(suffix_text)]}{suffix_text}"
        if candidate not in used_titles:
            used_titles.add(candidate)
            return candidate
        suffix += 1


def _style_header_row(
    worksheet,
    row_number: int,
    start_column: int = 1,
    end_column: int | None = None,
) -> None:
    end_column = end_column or worksheet.max_column
    for column in range(start_column, end_column + 1):
        cell = worksheet.cell(row=row_number, column=column)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def _autosize_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 36)


def _dedupe_preserve_order(values: list[str]) -> list[str]:
    return list(dict.fromkeys(value for value in values if value))


def _is_attention_row(row: dict[str, object]) -> bool:
    return bool(row.get("requires_attention"))


def _month_report_title(month: int, year: int) -> str:
    return f"Timesheet Report - {calendar.month_name[month]} {year}"


def _format_hours_cell(cell) -> None:
    cell.alignment = Alignment(horizontal="right")


def _write_row(worksheet, row_number: int, start_column: int, values: list[object]) -> None:
    for column_offset, value in enumerate(values):
        worksheet.cell(row=row_number, column=start_column + column_offset, value=value)


def _sop_code_columns() -> list[str]:
    configured_codes = [
        _normalize_note_code(sop_code.get("code"))
        for sop_code in load_sop_codes()
    ]
    return _dedupe_preserve_order(STANDARD_SOP_CODES + configured_codes)


def _worker_names_and_types(reviewed_rows: list[dict[str, object]]) -> dict[str, str]:
    workers = {
        str(worker.get("name", "")).strip(): str(worker.get("worker_type", "")).strip()
        for worker in load_workers()
        if str(worker.get("name", "")).strip()
    }
    for row in reviewed_rows:
        worker_name = str(row.get("worker", "")).strip() or "Unassigned"
        workers.setdefault(worker_name, "Unknown")
    return workers


def _merge_day_entry(existing: dict[str, object], row: dict[str, object]) -> None:
    note = str(row.get("notes", "") or "").strip()
    review_reason = str(row.get("review_reason", "") or "").strip()
    start_time = str(row.get("start_time", "") or "").strip()
    end_time = str(row.get("end_time", "") or "").strip()
    minutes = _parse_hours_to_minutes(row.get("hours_worked"))

    if start_time and not existing["start_time"]:
        existing["start_time"] = start_time
    if end_time:
        existing["end_time"] = end_time
    existing["minutes"] = int(existing["minutes"]) + minutes
    if note:
        existing_notes = list(existing["notes"])
        existing_notes.append(note)
        existing["notes"] = _dedupe_preserve_order(existing_notes)
    existing["requires_attention"] = bool(existing["requires_attention"]) or _is_attention_row(row)
    if _is_attention_row(row) and review_reason:
        existing_reasons = list(existing["review_reasons"])
        existing_reasons.append(review_reason)
        existing["review_reasons"] = _dedupe_preserve_order(existing_reasons)


def export_reviewed_rows_to_excel(
    session_id: str,
    reviewed_rows: list[dict[str, object]],
) -> Path:
    if not reviewed_rows:
        raise ValueError("Reviewed rows are required before export.")

    session = load_session(session_id)
    year = int(session["year"])
    month = int(session["month"])
    report_title = _month_report_title(month, year)
    export_path = get_session_exports_directory(session_id) / f"{report_title}.xlsx"
    last_day = calendar.monthrange(year, month)[1]
    sop_codes = _sop_code_columns()
    workers = _worker_names_and_types(reviewed_rows)

    rows_by_worker: dict[str, dict[int, dict[str, object]]] = defaultdict(dict)
    for row in reviewed_rows:
        worker = str(row.get("worker", "")).strip() or "Unassigned"
        work_date = _parse_work_date(row.get("date"))
        if not work_date or work_date.year != year or work_date.month != month:
            continue

        day = work_date.day
        rows_by_worker[worker].setdefault(
            day,
            {
                "start_time": "",
                "end_time": "",
                "minutes": 0,
                "notes": [],
                "requires_attention": False,
                "review_reasons": [],
            },
        )
        _merge_day_entry(rows_by_worker[worker][day], row)

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.sheet_view.showGridLines = False
    used_titles: set[str] = set()
    summary_rows = []

    for worker_name in sorted(workers):
        worksheet = workbook.create_sheet(_safe_sheet_title(worker_name, used_titles))
        worksheet.sheet_view.showGridLines = False
        total_minutes = 0
        note_counts: dict[str, int] = defaultdict(int)
        requires_attention = False
        days_worked = 0
        attention_items = []
        worker_type = workers.get(worker_name, "Unknown")
        worker_days = rows_by_worker.get(worker_name, {})

        table_header_row = 2
        table_start_column = 2
        table_end_column = 6
        _write_row(
            worksheet,
            table_header_row,
            table_start_column,
            ["Day", "Start Time", "End Time", "Total Time Worked", "Notes"],
        )
        _style_header_row(worksheet, table_header_row, table_start_column, table_end_column)

        for day in range(1, last_day + 1):
            entry = worker_days.get(
                day,
                {
                    "start_time": "",
                    "end_time": "",
                    "minutes": 0,
                    "notes": [],
                    "requires_attention": False,
                    "review_reasons": [],
                },
            )
            minutes = int(entry["minutes"])
            notes = [str(note).strip() for note in entry["notes"] if str(note).strip()]
            note_text = ", ".join(notes)
            review_reasons = [
                str(reason).strip()
                for reason in entry["review_reasons"]
                if str(reason).strip()
            ]
            total_minutes += minutes
            if minutes > 0:
                days_worked += 1
            for note in notes:
                note_code = _normalize_note_code(note)
                if note_code in sop_codes:
                    note_counts[note_code] += 1
            requires_attention = requires_attention or bool(entry["requires_attention"])
            if entry["requires_attention"]:
                attention_items.append(
                    {
                        "day": day,
                        "notes": note_text,
                        "review_reason": "; ".join(review_reasons),
                    }
                )

            current_row = table_header_row + day
            _write_row(
                worksheet,
                current_row,
                table_start_column,
                [
                    day,
                    str(entry["start_time"]),
                    str(entry["end_time"]),
                    _format_minutes(minutes),
                    note_text,
                ]
            )
            _format_hours_cell(worksheet.cell(row=current_row, column=5))
            if any(_normalize_note_code(note) in sop_codes for note in notes):
                for column in range(table_start_column, table_end_column + 1):
                    worksheet.cell(row=current_row, column=column).fill = NOTE_FILL
            for column in range(table_start_column, table_end_column + 1):
                worksheet.cell(row=current_row, column=column).border = THIN_BORDER

        total_hours = _format_minutes(total_minutes)
        total_row = worksheet.max_row + 2
        _write_row(
            worksheet,
            total_row,
            table_start_column,
            ["Total Hours Worked", total_hours, "", "", ""],
        )
        worksheet.cell(row=total_row, column=2).font = Font(bold=True)
        worksheet.cell(row=total_row, column=2).fill = SUMMARY_FILL
        worksheet.cell(row=total_row, column=3).fill = SUMMARY_FILL
        _format_hours_cell(worksheet.cell(row=total_row, column=3))
        for column in range(table_start_column, table_end_column + 1):
            worksheet.cell(row=total_row, column=column).border = THIN_BORDER

        for code in sop_codes:
            summary_row = worksheet.max_row + 1
            _write_row(
                worksheet,
                summary_row,
                table_start_column,
                [f"{code} Taken", note_counts.get(code, 0), "", "", ""],
            )
            worksheet.cell(row=summary_row, column=2).font = Font(bold=True)
            for column in range(table_start_column, table_end_column + 1):
                worksheet.cell(row=summary_row, column=column).border = THIN_BORDER

        if attention_items:
            attention_title_row = worksheet.max_row + 2
            attention_title_cell = worksheet.cell(
                row=attention_title_row,
                column=table_start_column,
                value="Attention Items",
            )
            attention_title_cell.font = Font(bold=True)

            attention_header_row = attention_title_row + 1
            attention_headers = ["Day", "Notes", "Review Reason"]
            attention_end_column = table_start_column + len(attention_headers) - 1
            _write_row(
                worksheet,
                attention_header_row,
                table_start_column,
                attention_headers,
            )
            _style_header_row(
                worksheet,
                attention_header_row,
                table_start_column,
                attention_end_column,
            )

            for attention_index, attention_item in enumerate(attention_items):
                attention_row = attention_header_row + attention_index + 1
                _write_row(
                    worksheet,
                    attention_row,
                    table_start_column,
                    [
                        attention_item["day"],
                        attention_item["notes"],
                        attention_item["review_reason"],
                    ],
                )
                for column in range(table_start_column, attention_end_column + 1):
                    worksheet.cell(row=attention_row, column=column).border = THIN_BORDER

        worksheet.freeze_panes = "B3"
        _autosize_columns(worksheet)

        summary_rows.append(
            {
                "worker_name": worker_name,
                "worker_type": worker_type,
                "days_worked": days_worked,
                "total_hours": total_hours,
                "note_counts": dict(note_counts),
                "requires_attention": "Yes" if requires_attention else "No",
                "total_minutes": total_minutes,
            }
        )

    summary_header_row = 2
    summary_start_column = 2
    summary_headers = [
        "Worker Name",
        "Worker Type",
        "Days Worked",
        "Total Hours Worked",
        *[f"{code} Taken" for code in sop_codes],
        "Requires Attention",
    ]
    summary_end_column = summary_start_column + len(summary_headers) - 1
    requires_attention_column = summary_end_column
    _write_row(summary_sheet, summary_header_row, summary_start_column, summary_headers)
    _style_header_row(
        summary_sheet,
        summary_header_row,
        summary_start_column,
        summary_end_column,
    )
    for row_index, row in enumerate(summary_rows):
        summary_row = summary_header_row + row_index + 1
        _write_row(
            summary_sheet,
            summary_row,
            summary_start_column,
            [
                row["worker_name"],
                row["worker_type"],
                row["days_worked"],
                row["total_hours"],
                *[row["note_counts"].get(code, 0) for code in sop_codes],
                row["requires_attention"],
            ]
        )
        row_fill = LIGHT_ORANGE_FILL if row_index % 2 == 0 else WHITE_FILL
        for column in range(summary_start_column, summary_end_column + 1):
            cell = summary_sheet.cell(row=summary_row, column=column)
            cell.fill = row_fill
            cell.border = THIN_BORDER
        _format_hours_cell(summary_sheet.cell(row=summary_row, column=5))
        if row["requires_attention"] == "Yes":
            summary_sheet.cell(row=summary_row, column=requires_attention_column).fill = ATTENTION_FILL
    _autosize_columns(summary_sheet)
    summary_sheet.freeze_panes = "B3"

    workbook.save(export_path)
    return export_path
